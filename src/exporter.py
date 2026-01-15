"""
数据导出模块

提供数据格式转换和文件导出功能
"""
import logging
from pathlib import Path
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from src.exceptions import FileError

logger = logging.getLogger(__name__)


class DataExporter:
    """数据导出器"""

    def export_to_excel(
        self, data: pd.DataFrame, output_path: Path, sheet_name: str = "Data"
    ) -> None:
        """
        导出数据到Excel文件

        Args:
            data: 要导出的DataFrame
            output_path: 输出文件路径
            sheet_name: 工作表名称

        Raises:
            FileError: 文件写入失败
        """
        try:
            logger.info(f"开始导出数据到Excel: {output_path}")

            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # 导出到Excel
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)

            logger.info(f"成功导出 {len(data)} 行数据到 {output_path}")

        except Exception as e:
            logger.error(f"导出Excel失败: {e}")
            raise FileError(f"无法导出Excel文件: {str(e)}")

    def export_to_csv(self, data: pd.DataFrame, output_path: Path) -> None:
        """
        导出数据到CSV文件

        Args:
            data: 要导出的DataFrame
            output_path: 输出文件路径

        Raises:
            FileError: 文件写入失败
        """
        try:
            logger.info(f"开始导出数据到CSV: {output_path}")

            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # 导出到CSV（使用UTF-8编码）
            data.to_csv(output_path, index=False, encoding="utf-8-sig")

            logger.info(f"成功导出 {len(data)} 行数据到 {output_path}")

        except Exception as e:
            logger.error(f"导出CSV失败: {e}")
            raise FileError(f"无法导出CSV文件: {str(e)}")

    def export_comparison_report(
        self, comparison_df: pd.DataFrame, output_path: Path
    ) -> None:
        """
        导出格式化的对比报告（Excel）

        Args:
            comparison_df: 对比数据DataFrame
            output_path: 输出文件路径

        特性：
        - 应用条件格式（增长=红色，下降=绿色）
        - 自动调整列宽
        - 添加标题和说明

        Raises:
            FileError: 文件写入失败
        """
        try:
            logger.info(f"开始导出格式化对比报告: {output_path}")

            # 确保输出目录存在
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "对比报告"

            # 定义样式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            center_alignment = Alignment(horizontal="center", vertical="center")

            # 增长样式（红色）
            increase_fill = PatternFill(
                start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"
            )
            # 下降样式（绿色）
            decrease_fill = PatternFill(
                start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"
            )
            # 无变化样式（灰色）
            no_change_fill = PatternFill(
                start_color="F0F0F0", end_color="F0F0F0", fill_type="solid"
            )

            # 写入标题行
            headers = list(comparison_df.columns)
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment

            # 写入数据行
            for row_idx, row_data in enumerate(
                dataframe_to_rows(comparison_df, index=False, header=False), start=2
            ):
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_alignment

                # 根据change_status列应用条件格式
                if "change_status" in headers:
                    status_col_idx = headers.index("change_status") + 1
                    status_value = ws.cell(row=row_idx, column=status_col_idx).value

                    # 为整行应用颜色
                    if status_value == "increase":
                        for col_idx in range(1, len(headers) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = increase_fill
                    elif status_value == "decrease":
                        for col_idx in range(1, len(headers) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = decrease_fill
                    elif status_value == "no_change":
                        for col_idx in range(1, len(headers) + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = no_change_fill

            # 自动调整列宽
            for col_idx, column in enumerate(ws.columns, start=1):
                max_length = 0
                column_letter = ws.cell(row=1, column=col_idx).column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)  # 最大宽度50
                ws.column_dimensions[column_letter].width = adjusted_width

            # 保存文件
            wb.save(output_path)

            logger.info(f"成功导出格式化对比报告到 {output_path}")

        except Exception as e:
            logger.error(f"导出对比报告失败: {e}")
            raise FileError(f"无法导出对比报告: {str(e)}")
