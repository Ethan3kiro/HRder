"""
数据导出模块的单元测试

测试DataExporter类的各项功能
"""
import pytest
import pandas as pd
from pathlib import Path
import tempfile
from openpyxl import load_workbook

from src.exporter import DataExporter
from src.exceptions import FileError


class TestDataExporter:
    """DataExporter类的单元测试"""

    def test_export_to_excel_basic(self):
        """测试基本的Excel导出功能"""
        # 准备测试数据
        data = pd.DataFrame(
            {
                "item_name": ["Forward Power", "PA Voltage", "Ambient Temp"],
                "value": [450.5, 12.3, 28.5],
                "unit": ["W", "V", "°C"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_output.xlsx"

            # 导出数据
            exporter.export_to_excel(data, output_path, sheet_name="TestSheet")

            # 验证文件存在
            assert output_path.exists()

            # 读取并验证数据
            read_data = pd.read_excel(output_path, sheet_name="TestSheet")
            assert len(read_data) == 3
            assert list(read_data.columns) == ["item_name", "value", "unit"]
            assert read_data["item_name"].tolist() == [
                "Forward Power",
                "PA Voltage",
                "Ambient Temp",
            ]
            assert read_data["value"].tolist() == [450.5, 12.3, 28.5]

    def test_export_to_excel_empty_dataframe(self):
        """测试导出空DataFrame"""
        data = pd.DataFrame(columns=["item_name", "value", "unit"])

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "empty_output.xlsx"

            # 导出空数据
            exporter.export_to_excel(data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 读取并验证
            read_data = pd.read_excel(output_path)
            assert len(read_data) == 0
            assert list(read_data.columns) == ["item_name", "value", "unit"]

    def test_export_to_excel_creates_directory(self):
        """测试导出时自动创建目录"""
        data = pd.DataFrame(
            {"item_name": ["Test Item"], "value": [100.0], "unit": ["V"]}
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            # 使用不存在的子目录
            output_path = Path(tmpdir) / "subdir" / "nested" / "output.xlsx"

            # 导出数据
            exporter.export_to_excel(data, output_path)

            # 验证文件和目录都被创建
            assert output_path.exists()
            assert output_path.parent.exists()

    def test_export_to_csv_basic(self):
        """测试基本的CSV导出功能"""
        # 准备测试数据
        data = pd.DataFrame(
            {
                "item_name": ["Reflected Power", "PA Current", "Airflow %"],
                "value": [25.3, 8.7, 85.0],
                "unit": ["W", "A", "%"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "test_output.csv"

            # 导出数据
            exporter.export_to_csv(data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 读取并验证数据
            read_data = pd.read_csv(output_path, encoding="utf-8-sig")
            assert len(read_data) == 3
            assert list(read_data.columns) == ["item_name", "value", "unit"]
            assert read_data["item_name"].tolist() == [
                "Reflected Power",
                "PA Current",
                "Airflow %",
            ]
            assert read_data["value"].tolist() == [25.3, 8.7, 85.0]

    def test_export_to_csv_with_chinese_characters(self):
        """测试CSV导出中文字符"""
        # 准备包含中文的测试数据
        data = pd.DataFrame(
            {
                "item_name": ["前向功率", "反射功率", "环境温度"],
                "value": [450.5, 25.3, 28.5],
                "unit": ["W", "W", "°C"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "chinese_output.csv"

            # 导出数据
            exporter.export_to_csv(data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 读取并验证数据（使用UTF-8-sig编码）
            read_data = pd.read_csv(output_path, encoding="utf-8-sig")
            assert len(read_data) == 3
            assert read_data["item_name"].tolist() == ["前向功率", "反射功率", "环境温度"]

    def test_export_comparison_report_with_formatting(self):
        """测试格式化对比报告导出"""
        # 准备对比数据
        comparison_data = pd.DataFrame(
            {
                "item_name": [
                    "Forward Power",
                    "PA Voltage",
                    "Ambient Temp",
                    "PA Current",
                ],
                "value_month1": [450.0, 12.0, 28.0, 8.5],
                "value_month2": [480.0, 11.5, 29.0, 8.5],
                "unit": ["W", "V", "°C", "A"],
                "absolute_change": [30.0, -0.5, 1.0, 0.0],
                "relative_change": [6.67, -4.17, 3.57, 0.0],
                "change_status": ["increase", "decrease", "increase", "no_change"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "comparison_report.xlsx"

            # 导出对比报告
            exporter.export_comparison_report(comparison_data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 读取并验证数据
            read_data = pd.read_excel(output_path)
            assert len(read_data) == 4
            assert list(read_data.columns) == list(comparison_data.columns)

            # 验证数据内容
            assert (
                read_data["item_name"].tolist() == comparison_data["item_name"].tolist()
            )
            assert read_data["change_status"].tolist() == [
                "increase",
                "decrease",
                "increase",
                "no_change",
            ]

            # 验证格式（通过openpyxl检查）
            wb = load_workbook(output_path)
            ws = wb.active

            # 验证工作表名称
            assert ws.title == "对比报告"

            # 验证有标题行
            assert ws.cell(row=1, column=1).value == "item_name"

            # 验证数据行存在
            assert ws.cell(row=2, column=1).value == "Forward Power"

    def test_export_comparison_report_applies_colors(self):
        """测试对比报告应用颜色格式"""
        # 准备包含不同状态的数据
        comparison_data = pd.DataFrame(
            {
                "item_name": ["Item1", "Item2", "Item3"],
                "value_month1": [100.0, 200.0, 300.0],
                "value_month2": [120.0, 180.0, 300.0],
                "unit": ["V", "V", "V"],
                "absolute_change": [20.0, -20.0, 0.0],
                "relative_change": [20.0, -10.0, 0.0],
                "change_status": ["increase", "decrease", "no_change"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "colored_report.xlsx"

            # 导出对比报告
            exporter.export_comparison_report(comparison_data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 使用openpyxl验证格式
            wb = load_workbook(output_path)
            ws = wb.active

            # 验证增长行（第2行）有红色背景
            increase_cell = ws.cell(row=2, column=1)
            assert increase_cell.fill.start_color.rgb in ["FFFFCCCC", "00FFCCCC"]  # 红色

            # 验证下降行（第3行）有绿色背景
            decrease_cell = ws.cell(row=3, column=1)
            assert decrease_cell.fill.start_color.rgb in ["FFCCFFCC", "00CCFFCC"]  # 绿色

            # 验证无变化行（第4行）有灰色背景
            no_change_cell = ws.cell(row=4, column=1)
            assert no_change_cell.fill.start_color.rgb in ["FFF0F0F0", "00F0F0F0"]  # 灰色

    def test_export_comparison_report_auto_adjusts_column_width(self):
        """测试对比报告自动调整列宽"""
        # 准备包含长文本的数据
        comparison_data = pd.DataFrame(
            {
                "item_name": ["Very Long Item Name That Should Adjust Column Width"],
                "value_month1": [100.0],
                "value_month2": [120.0],
                "unit": ["V"],
                "absolute_change": [20.0],
                "relative_change": [20.0],
                "change_status": ["increase"],
            }
        )

        exporter = DataExporter()

        with tempfile.TemporaryDirectory() as tmpdir:
            output_path = Path(tmpdir) / "wide_report.xlsx"

            # 导出对比报告
            exporter.export_comparison_report(comparison_data, output_path)

            # 验证文件存在
            assert output_path.exists()

            # 使用openpyxl验证列宽
            wb = load_workbook(output_path)
            ws = wb.active

            # 验证第一列（item_name）的宽度被调整
            column_width = ws.column_dimensions["A"].width
            assert column_width > 10  # 应该比默认宽度大
            assert column_width <= 50  # 但不应超过最大宽度

    def test_export_to_excel_invalid_path_raises_error(self):
        """测试无效路径导出Excel抛出错误"""
        data = pd.DataFrame({"item_name": ["Test"], "value": [100.0], "unit": ["V"]})

        exporter = DataExporter()

        # 使用无效路径（只读文件系统路径）
        invalid_path = Path("/invalid/readonly/path/output.xlsx")

        # 应该抛出FileError
        with pytest.raises(FileError):
            exporter.export_to_excel(data, invalid_path)

    def test_export_to_csv_invalid_path_raises_error(self):
        """测试无效路径导出CSV抛出错误"""
        data = pd.DataFrame({"item_name": ["Test"], "value": [100.0], "unit": ["V"]})

        exporter = DataExporter()

        # 使用无效路径
        invalid_path = Path("/invalid/readonly/path/output.csv")

        # 应该抛出FileError
        with pytest.raises(FileError):
            exporter.export_to_csv(data, invalid_path)

    def test_export_comparison_report_invalid_path_raises_error(self):
        """测试无效路径导出对比报告抛出错误"""
        comparison_data = pd.DataFrame(
            {
                "item_name": ["Test"],
                "value_month1": [100.0],
                "value_month2": [120.0],
                "unit": ["V"],
                "absolute_change": [20.0],
                "relative_change": [20.0],
                "change_status": ["increase"],
            }
        )

        exporter = DataExporter()

        # 使用无效路径
        invalid_path = Path("/invalid/readonly/path/report.xlsx")

        # 应该抛出FileError
        with pytest.raises(FileError):
            exporter.export_comparison_report(comparison_data, invalid_path)
